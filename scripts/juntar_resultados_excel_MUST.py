import re
import pandas as pd
from openpyxl import load_workbook
import os

def substituir_aba_excel(df_novo, caminho_arquivo, nome_aba, engine='openpyxl'):
    """
    Substitui uma aba específica em um arquivo Excel existente por um novo DataFrame.
    
    Args:
        df_novo: DataFrame do pandas que substituirá a aba existente
        caminho_arquivo: Caminho completo do arquivo Excel
        nome_aba: Nome da aba a ser substituída
        engine: Motor do Excel a ser usado (padrão: 'openpyxl')
    """
    # Verifica se o arquivo existe
    if not os.path.exists(caminho_arquivo):
        print(f"⚠️ Arquivo não encontrado: {caminho_arquivo}")
        return False
    
    try:
        # Carrega o workbook existente
        book = load_workbook(caminho_arquivo)
        
        # Remove a aba se já existir
        if nome_aba in book.sheetnames:
            del book[nome_aba]
        
        # Salva as alterações temporárias
        book.save(caminho_arquivo)
        book.close()
        
        # Adiciona o novo DataFrame na aba especificada
        with pd.ExcelWriter(caminho_arquivo, engine=engine, mode='a') as writer:
            df_novo.to_excel(writer, sheet_name=nome_aba, index=False)
        
        print(f"\n✅ Aba '{nome_aba}' substituída com sucesso em {caminho_arquivo}")
        return True
        
    except Exception as e:
        print(f"❌ Erro ao substituir aba: {e}")
        return False

def extrair_empresa(nome_arquivo: str) -> str:
    """
    Extrai o nome da empresa a partir do nome do arquivo.
    Exemplos:
        "saida_anotacoes_CUST-2002-025-41 - SUL SUDESTE_minuta_recon_2025_28_final.xlsx" -> "SUL SUDESTE"
        "saida_anotacoes_CUST-2002-114-64_ CPFL_Paulista.xlsx" -> "CPFL PAULISTA"
    """
    # Remove prefixo e extensão
    base = os.path.splitext(nome_arquivo)[0]
    base = re.sub(r"^saida_anotacoes_", "", base)

    # Padrão para encontrar o código (ex: CUST-2002-114-64)
    padrao_codigo = r'^CUST-\d{4}-\d{2,4}-\d{2,3}'
    match = re.search(padrao_codigo, base)
    
    if not match:
        return "DESCONHECIDA"
    
    # Remove o código e separadores subsequentes
    empresa_raw = base[match.end():]
    empresa_raw = re.sub(r"^[\s_-]+", "", empresa_raw)  # Remove hífens/underscores iniciais

    # Remove termos indesejados e tudo após eles
    empresa_limpa = re.split(
        r"[\s_-]*(?:minuta|recon|final|202[0-9])[\s_-]*", 
        empresa_raw, 
        flags=re.IGNORECASE
    )[0]

    # Remove caracteres especiais e espaços extras
    empresa_limpa = re.sub(r"[_\s]+", " ", empresa_limpa).strip()
    return empresa_limpa.upper() if empresa_limpa else "DESCONHECIDA"

def consolidar_anotacoes(diretorio: str):
    """
    Consolida os arquivos de anotações exportados em um único Excel.
    - Só concatena arquivos com colunas iguais.
    - Filtra num_tabela = 1 (quando existir).
    - Extrai nome da empresa do arquivo.
    """
    arquivos = [f for f in os.listdir(diretorio) if f.endswith(".xlsx") and f.startswith("saida_anotacoes")]
    
    if not arquivos:
        print("⚠️ Nenhum arquivo encontrado para consolidar.")
        return
    
    dataframes = []
    empresas = set()
    colunas_padrao = None

    for arq in arquivos:
        caminho = os.path.join(diretorio, arq)
        try:
            df = pd.read_excel(caminho)

            # Filtra apenas num_tabela = 1
            if "num_tabela" in df.columns:
                df = df[df["num_tabela"] == 1]

            # Garante colunas iguais
            if colunas_padrao is None:
                colunas_padrao = list(df.columns)
                print(f"📊 Colunas padrão definidas a partir de {arq}: {colunas_padrao}")
            elif list(df.columns) != colunas_padrao:
                print(f"⏭️ Ignorando {arq} pois as colunas não batem com o padrão.")
                continue

            # Extrai empresa
            empresa = extrair_empresa(arq)
            df.insert(0, "EMPRESA", empresa)  # força ser a primeira coluna
            df["Arquivo_Origem"] = arq
            empresas.add(empresa)

            dataframes.append(df)

        except Exception as e:
            print(f"❌ Erro ao ler {arq}: {e}")

    if not dataframes:
        print("⚠️ Nenhum dado válido para consolidar.")
        return

    df_final = pd.concat(dataframes, ignore_index=True)

    # Exporta para Excel
    caminho_saida = os.path.join(diretorio, "export_notes_MUST_tables.xlsx")
    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df_final.to_excel(writer, sheet_name="Notas Consolidada", index=False)
        pd.DataFrame({"Empresas": sorted(empresas)}).to_excel(writer, sheet_name="Empresas", index=False)

    print(f"✅ Consolidação concluída: {caminho_saida}")
    print(f"🔎 {len(empresas)} empresas identificadas: {sorted(empresas)}")

# -----------------------------
# Função para limpar código ONS
# -----------------------------
def extrair_cod_ons(valor):
    if pd.isna(valor):
        return None
    texto = str(valor).strip().upper()
    # Captura prefixo + dígito opcional (SPXXXX-138 ou SPXXXX138)
    match = re.search(r"([A-Z]{2,}[A-Z0-9]*)(?:\s*-?\s*(\d{2,3}))?", texto)
    if not match:
        return texto
    prefixo = match.group(1)  # letras + dígitos (SPASS, SPUFA, SPBRB...)
    sufixo = match.group(2)   # tensão (ex: 138, 88, etc)
    return f"{prefixo}-{sufixo}" if sufixo else prefixo

def normalizar_cod_ons(valor):
    if pd.isna(valor):
        return None
    texto = str(valor).upper().strip()
    texto = re.sub(r"\s+", "", texto)  # remove espaços internos
    
    # Detecta padrões do tipo PREFIXO + NUMERO
    match = re.match(r"([A-Z]{2,}[A-Z0-9]*?)(\d{2,4})$", texto)
    if match:
        prefixo, sufixo = match.groups()
        sufixo = str(int(sufixo))  # remove zeros à esquerda
        return f"{prefixo}-{sufixo}"
    
    # Se já tiver hífen, normaliza
    match = re.match(r"([A-Z]{2,}[A-Z0-9]*)-?(\d{2,4})$", texto)
    if match:
        prefixo, sufixo = match.groups()
        sufixo = str(int(sufixo))  # remove zeros à esquerda
        return f"{prefixo}-{sufixo}"
    
    return texto

# -----------------------------


def juntar_resultados_tabelas_MUST():

    # Caminho da pasta onde estão os arquivos de anotações
    diretorio_anotacoes = r"C:\Users\pedrovictor.veras\OneDrive - Operador Nacional do Sistema Eletrico\Documentos\ESTAGIO_ONS_PVRV_2025\AUTOMACÕES ONS\arquivos\anotacoes_extraidas"

    # Executar consolidação de anotações
    consolidar_anotacoes(diretorio_anotacoes)

    # -----------------------------
    # Caminho do arquivo Excel
    # -----------------------------
    path = r"C:\Users\pedrovictor.veras\OneDrive - Operador Nacional do Sistema Eletrico\Documentos\ESTAGIO_ONS_PVRV_2025\AUTOMACÕES ONS\arquivos\database\PROTOTIPO_database.xlsx"
    path = r"C:\Users\pedrovictor.veras\OneDrive - Operador Nacional do Sistema Eletrico\Documentos\ESTAGIO_ONS_PVRV_2025\AUTOMACÕES ONS\arquivos\tabelas_extraidas\database_must.xlsx"

    # -----------------------------
    # Leitura das planilhas
    # -----------------------------
    planilha_must = pd.read_excel(path, sheet_name="Tabelas Consolidada")

    # Exibe as primeiras linhas da planilha MUST
    #df_tables = pd.read_excel(path, sheet_name="ANOTAÇÕES")

    # Carrega a planilha MUST
    df_notes = pd.read_excel(r"C:\Users\pedrovictor.veras\OneDrive - Operador Nacional do Sistema Eletrico\Documentos\ESTAGIO_ONS_PVRV_2025\AUTOMACÕES ONS\arquivos\anotacoes_extraidas\export_notes_MUST_tables.xlsx")


    # Quando estiver pronto para substituir a aba:
    substituir_aba_excel(df_notes, path, "TABELAS")

    #  Verificar se foi apenas a tabela 1 no excel database
    print(df_notes.head(5))
    print(df_notes.shape)
    print(df_notes["EMPRESA"].value_counts())




    # -----------------------------
    # Aplicando a limpeza nos códigos ONS
    # -----------------------------
    # Padronizar códigos ONS
    planilha_must["Cód ONS"] = planilha_must["Cód ONS"].apply(extrair_cod_ons).str.upper().str.strip()
    df_notes["Cód ONS"] = df_notes["Cód ONS"].apply(extrair_cod_ons).str.upper().str.strip()

    #! TODO: revisar se é necessário normalizar mais
    # Aplica normalização ao COD ONS
    #planilha_must["Cód ONS"] = planilha_must["Cód ONS"].apply(normalizar_cod_ons)
    #df_notes["Cód ONS"] = df_notes["Cód ONS"].apply(normalizar_cod_ons)


    print("\nPrimeiras linhas das anotações após limpeza do código ONS:")
    df_notes_filtrado = df_notes[df_notes["Num_Tabela"] == 1].reset_index(drop=True) # Filtra apenas as anotações da Tabela 1
    print(df_notes_filtrado.shape)
    print(df_notes_filtrado.head(5))

    # -----------------------------
    # Merge das tabelas com as anotações
    # -----------------------------
    tabela = planilha_must.merge(
        df_notes_filtrado[["Cód ONS", "Anotacao"]],
        on="Cód ONS",
        how="left"
    )

    # Exibe resultado final
    print("\n\nTabela MUST consolidada:")
    #tabela = tabela[tabela["num_tabela"] == 1].reset_index(drop=True) # Filtra apenas as anotações da Tabela 1

    print(tabela.shape)
    #print(tabela.columns)
    print(tabela)



    tabela.to_excel(
        r"C:\Users\pedrovictor.veras\OneDrive - Operador Nacional do Sistema Eletrico\Documentos\ESTAGIO_ONS_PVRV_2025\AUTOMACÕES ONS\arquivos\database\must_tables_PDF_notes_merged.xlsx",
        index=False
    )



    tabela.to_json(
        r"C:\Users\pedrovictor.veras\OneDrive - Operador Nacional do Sistema Eletrico\Documentos\ESTAGIO_ONS_PVRV_2025\GitHub\dashboard-website-template\dashboard_must_webiste\must_tables_PDF_notes_merged.json",
        orient="records",
        force_ascii=False
    )
