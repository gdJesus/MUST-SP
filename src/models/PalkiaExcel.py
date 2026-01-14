# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re

# Certifique-se de ter as bibliotecas necessárias instaladas:
# pip install pandas openpyxl

class Logger:
    """Classe simplificada para logs claros no console."""
    def log(self, message, status="INFO"):
        print(f"[{status}] {message}")

class DatabaseBuilder:
    """
    Classe responsável por transformar um DataFrame de dados extraídos (formato longo)
    em um formato de banco de dados (formato largo), usando 'Cód ONS' como chave primária.
    """
    def __init__(self):
        self.logger = Logger()

    def transform_to_database_format(self, long_df: pd.DataFrame) -> pd.DataFrame:
        """
        Converte o DataFrame extraído para o formato de banco de dados.
        Cada 'Cód ONS' se torna uma linha única.

        Args:
            long_df (pd.DataFrame): O DataFrame extraído, com múltiplas linhas por Cód ONS.

        Returns:
            pd.DataFrame: Um novo DataFrame no formato de banco de dados.
        """
        if long_df.empty:
            self.logger.log("DataFrame de entrada está vazio. Nenhuma transformação a ser feita.", "WARNING")
            return pd.DataFrame()

        self.logger.log("Iniciando transformação para o formato de banco de dados...", "INFO")
        
        # --- GUIA DE ESTUDO: MODELAGEM DOS DADOS ---

        # 1. DEFINIR COLUNAS-CHAVE (ID):
        #    Estas são as colunas que identificam uma instalação de forma única,
        #    independentemente do período de tempo. Elas permanecerão como colunas fixas.
        id_cols = ['Cód ONS', 'Instalação', 'Tensão (kV)']
        
        # 2. DEFINIR COLUNAS DE VALORES (A SEREM TRANSFORMADAS):
        #    Estas são as colunas cujos valores dependem do período ('De' e 'Até').
        #    O script irá "desempilhar" (pivotar) estas colunas.
        value_cols = [col for col in long_df.columns if col not in id_cols + ['De', 'Até']]

        db_records = []
        # 3. AGRUPAR PELA CHAVE PRIMÁRIA:
        #    Agrupamos o DataFrame por 'Cód ONS'. Isso nos dá um "pacote" de linhas
        #    para cada instalação, que vamos consolidar em uma única linha.
        for cod_ons, group in long_df.groupby('Cód ONS'):
            # 4. CRIAR O REGISTRO BASE:
            #    Para cada Cód ONS, criamos um dicionário inicial com as informações
            #    que são constantes para aquela instalação.
            record = {
                'Cód ONS': cod_ons,
                'Instalação': group['Instalação'].iloc[0], # Pega o valor da primeira linha do grupo
                'Tensão (kV)': group['Tensão (kV)'].iloc[0]
            }
            
            # 5. ITERAR E "PIVOTAR" OS PERÍODOS:
            #    Agora, iteramos sobre cada linha do grupo (cada período diferente).
            for _, row in group.iterrows():
                # Cria um sufixo único para as novas colunas baseado no período.
                # Ex: "_1/Jan_30/Jun"
                period_suffix = f"_{row['De']}_{row['Até']}"
                
                # Para cada coluna de valor, criamos uma nova coluna no nosso registro.
                for col in value_cols:
                    # Cria o novo nome da coluna. Ex: "Ponta 2025_Valor_1/Jan_30/Jun"
                    new_col_name = f"{col}{period_suffix}"
                    # Adiciona o valor ao nosso registro. .get() evita erros se a coluna não existir.
                    record[new_col_name] = row.get(col, '')
            
            db_records.append(record)
            
        self.logger.log(f"Transformação concluída. {len(db_records)} registros únicos criados.", "SUCCESS")
        # 6. CRIAR O DATAFRAME FINAL:
        #    Converte a lista de dicionários (um para cada Cód ONS) em um DataFrame final.
        return pd.DataFrame(db_records)

class PalkiaExcel:
    """
    Classe para gerenciar um arquivo Excel como um banco de dados,
    onde cada aba representa uma tabela de uma empresa. Utiliza openpyxl.
    """
    def __init__(self, filepath: str):
        """
        Inicializa o gerenciador do banco de dados Excel.

        Args:
            filepath (str): O caminho para o arquivo .xlsx (ex: "db.xlsx").
        """
        self.filepath = filepath
        self.logger = Logger()
        try:
            self.workbook = openpyxl.load_workbook(self.filepath)
            self.logger.log(f"Banco de dados '{os.path.basename(filepath)}' carregado.", "INFO")
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])
            self.logger.log(f"Novo banco de dados '{os.path.basename(filepath)}' criado.", "INFO")

    def update_sheet_with_data(self, sheet_name: str, df: pd.DataFrame):
        """
        Adiciona ou atualiza uma aba com os dados de um DataFrame.
        Se a aba já existir, ela será limpa e reescrita para garantir dados atualizados.
        """
        if df.empty:
            self.logger.log(f"DataFrame para a aba '{sheet_name}' está vazio. Pulando.", "WARNING")
            return

        safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '', sheet_name)[:31]

        if safe_sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[safe_sheet_name])
            self.logger.log(f"Aba '{safe_sheet_name}' existente foi removida para atualização.", "INFO")
        
        sheet = self.workbook.create_sheet(title=safe_sheet_name)
        
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)
        
        self.logger.log(f"{df.shape[0]} linhas de dados inseridas na aba '{safe_sheet_name}'.", "SUCCESS")

    def save(self):
        """
        Salva todas as alterações no arquivo Excel.
        """
        try:
            self.workbook.save(self.filepath)
            self.logger.log(f"Banco de dados salvo com sucesso em: {self.filepath}", "SUCCESS")
        except Exception as e:
            self.logger.log(f"Erro ao salvar o arquivo Excel: {e}", "ERROR")

# --- EXEMPLO DE USO ---

def get_company_name_from_filename(filename: str) -> str:
    """Função de exemplo para extrair o nome da empresa do nome do arquivo."""
    name = os.path.basename(filename)
    name = re.sub(r'^saida_CUST-\d{4}-\d{3}-\d{2}\s*-\s*', '', name, flags=re.IGNORECASE)
    name = os.path.splitext(name)[0]
    name = re.sub(r'\s*RECON.*', '', name, flags=re.IGNORECASE).strip()
    return name

if __name__ == "__main__":
    # --- Configuração do Processo ---
    extracted_data_folder = r"C:\Users\pedrovictor.veras\OneDrive - Operador Nacional do Sistema Eletrico\Documentos\ESTAGIO_ONS_PVRV_2025\AUTOMACÕES ONS\arquivos\tabelas_extraidas"
    db_output_folder = os.path.join(extracted_data_folder, "database_final")
    os.makedirs(db_output_folder, exist_ok=True)
    
    files_to_process = [
        "saida_CUST-2002-123-41 - JAGUARI - RECON 2025-2028.xlsx"
        # Adicione outros arquivos .xlsx ou .csv aqui
    ]
    
    db_filepath = os.path.join(db_output_folder, "db.xlsx")

    # 1. Inicializa as classes
    db_excel = PalkiaExcel(filepath=db_filepath)
    db_builder = DatabaseBuilder()

    # --- Loop de Processamento ---
    for filename in files_to_process:
        filepath = os.path.join(extracted_data_folder, filename)
        
        if not os.path.exists(filepath):
            print(f"AVISO: Arquivo de dados '{filename}' não encontrado, pulando.")
            continue
        
        print(f"\n--- Processando arquivo: {filename} ---")
        
        # 2. Lê os dados do arquivo Excel/CSV extraído
        try:
            if filename.endswith('.xlsx'):
                extracted_df = pd.read_excel(filepath)
            else:
                extracted_df = pd.read_csv(filepath)
        except Exception as e:
            print(f"ERRO ao ler o arquivo '{filename}': {e}")
            continue

        # 3. Transforma os dados para o formato de banco de dados
        db_df = db_builder.transform_to_database_format(extracted_df)
        
        # 4. Usa a PalkiaExcel para inserir os dados na aba correta
        if not db_df.empty:
            company_name = get_company_name_from_filename(filename)
            db_excel.update_sheet_with_data(company_name, db_df)

    # 5. Salva o arquivo Excel com todas as abas atualizadas
    db_excel.save()
