# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re
import sqlite3

# Certifique-se de ter as bibliotecas necessárias instaladas:
# pip install pandas openpyxl

class Logger:
    """Classe simplificada para logs claros no console."""
    def log(self, message, status="INFO"):
        print(f"[{status}] {message}")

class DatabaseBuilder:
    """
    Classe responsável por transformar um DataFrame de dados extraídos (formato longo)
    em um formato de banco de dados (formato largo), usando 'Cód ONS' como chave primária.
    """
    def __init__(self):
        self.logger = Logger()

    def transform_to_database_format(self, long_df: pd.DataFrame) -> pd.DataFrame:
        """
        Converte o DataFrame extraído para o formato de banco de dados.
        Cada 'Cód ONS' se torna uma linha única.

        Args:
            long_df (pd.DataFrame): O DataFrame extraído, com múltiplas linhas por Cód ONS.

        Returns:
            pd.DataFrame: Um novo DataFrame no formato de banco de dados.
        """
        if long_df.empty:
            self.logger.log("DataFrame de entrada está vazio. Nenhuma transformação a ser feita.", "WARNING")
            return pd.DataFrame()

        self.logger.log("Iniciando transformação para o formato de banco de dados...", "INFO")
        
        # Colunas que identificam unicamente uma instalação, sem o período.
        # Estas serão as colunas fixas no nosso novo banco de dados.
        id_cols = ['Cód ONS', 'Instalação', 'Tensão (kV)']
        
        # Colunas cujos valores mudarão com base no período (De/Até).
        # Estas colunas serão "desempilhadas".
        value_cols = [col for col in long_df.columns if col not in id_cols + ['De', 'Até']]

        db_records = []
        # Agrupamos o DataFrame por 'Cód ONS' para processar cada instalação unicamente.
        for cod_ons, group in long_df.groupby('Cód ONS'):
            # O registro base contém as informações que não mudam para um mesmo Cód ONS.
            record = {
                'Cód ONS': cod_ons,
                'Instalação': group['Instalação'].iloc[0],
                'Tensão (kV)': group['Tensão (kV)'].iloc[0]
            }
            
            # Itera sobre cada período diferente (cada linha dentro do grupo).
            for _, row in group.iterrows():
                # Cria um sufixo único para as novas colunas baseado no período.
                # Ex: "_1/Jan_30/Jun"
                period_suffix = f"_{row['De']}_{row['Até']}"
                
                # Adiciona os valores de MUST com o sufixo do período ao nosso registro.
                for col in value_cols:
                    # Cria o novo nome da coluna. Ex: "Ponta 2025_Valor_1/Jan_30/Jun"
                    new_col_name = f"{col}{period_suffix}"
                    # Adiciona o valor ao nosso registro. .get() evita erros se a coluna não existir.
                    record[new_col_name] = row.get(col, '')
            
            db_records.append(record)
            
        self.logger.log(f"Transformação concluída. {len(db_records)} registros únicos criados.", "SUCCESS")
        # Converte a lista de dicionários em um DataFrame final.
        return pd.DataFrame(db_records)

class PalkiaExcel:
    """
    Classe para gerenciar um arquivo Excel como um banco de dados,
    onde cada aba representa uma tabela de uma empresa. Utiliza openpyxl.
    """
    def __init__(self, filepath: str):
        """
        Inicializa o gerenciador do banco de dados Excel.

        Args:
            filepath (str): O caminho para o arquivo .xlsx (ex: "db.xlsx").
        """
        self.filepath = filepath
        self.logger = Logger()
        try:
            # Tenta carregar o arquivo Excel se ele já existir
            self.workbook = openpyxl.load_workbook(self.filepath)
            self.logger.log(f"Banco de dados '{os.path.basename(filepath)}' carregado.", "INFO")
        except FileNotFoundError:
            # Se não existir, cria um novo
            self.workbook = openpyxl.Workbook()
            # Remove a aba padrão "Sheet" que é criada automaticamente
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])
            self.logger.log(f"Novo banco de dados '{os.path.basename(filepath)}' criado.", "INFO")

    def update_sheet_with_data(self, sheet_name: str, df: pd.DataFrame):
        """
        Adiciona ou atualiza uma aba com os dados de um DataFrame.
        Se a aba já existir, ela será limpa e reescrita para garantir dados atualizados.

        Args:
            sheet_name (str): O nome da aba (ex: nome da empresa).
            df (pd.DataFrame): O DataFrame do pandas com os dados a serem inseridos.
        """
        if df.empty:
            self.logger.log(f"DataFrame para a aba '{sheet_name}' está vazio. Pulando.", "WARNING")
            return

        # Garante que o nome da aba seja válido para o Excel (máx 31 caracteres, sem caracteres inválidos)
        safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '', sheet_name)[:31]

        if safe_sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[safe_sheet_name])
            self.logger.log(f"Aba '{safe_sheet_name}' existente foi removida para atualização.", "INFO")
        
        sheet = self.workbook.create_sheet(title=safe_sheet_name)
        
        # Converte o DataFrame para linhas que o openpyxl pode usar e as escreve na planilha
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)
        
        self.logger.log(f"{df.shape[0]} linhas de dados inseridas na aba '{safe_sheet_name}'.", "SUCCESS")

    def save(self):
        """
        Salva todas as alterações no arquivo Excel.
        """
        try:
            self.workbook.save(self.filepath)
            self.logger.log(f"Banco de dados salvo com sucesso em: {self.filepath}", "SUCCESS")
        except Exception as e:
            self.logger.log(f"Erro ao salvar o arquivo Excel: {e}", "ERROR")

    def excel_to_sql(self, excel_path: str, db_path: str):
        """
        Importa dados do arquivo Excel para o banco de dados SQLite.

        Args:
            excel_path (str): Caminho para o arquivo Excel.
            db_path (str): Caminho para o banco de dados SQLite.
        """
        self.logger.log("Iniciando importação do Excel para o banco de dados SQLite...", "INFO")
        try:
            conn = sqlite3.connect(db_path)
            xls = pd.ExcelFile(excel_path)

            for sheet_name in xls.sheet_names:
                df = xls.parse(sheet_name)
                df.to_sql(sheet_name, conn, if_exists='replace', index=False)
                self.logger.log(f"Aba '{sheet_name}' importada com sucesso para o banco de dados.", "SUCCESS")

            conn.close()
        except Exception as e:
            self.logger.log(f"Erro ao importar Excel para SQLite: {e}", "ERROR")

    def sql_to_excel(self, db_path: str, excel_path: str):
        """
        Exporta dados do banco de dados SQLite para o arquivo Excel.

        Args:
            db_path (str): Caminho para o banco de dados SQLite.
            excel_path (str): Caminho para o arquivo Excel.
        """
        self.logger.log("Iniciando exportação do banco de dados SQLite para o Excel...", "INFO")
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()

            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                table_names = [name[0] for name in cursor.execute("SELECT name FROM sqlite_master WHERE type='table';").fetchall()]
                for table_name in table_names:
                    df = pd.read_sql_query(f"SELECT * FROM '{table_name}'", conn)
                    df.to_excel(writer, sheet_name=table_name, index=False)
                    self.logger.log(f"Tabela '{table_name}' exportada com sucesso para o Excel.", "SUCCESS")

            conn.close()
        except Exception as e:
            self.logger.log(f"Erro ao exportar SQLite para Excel: {e}", "ERROR")

class DatabaseOrchestrator:
    """
    Orquestra a criação do banco de dados Excel a partir de múltiplos arquivos
    de dados extraídos de diferentes pastas.
    """
    def __init__(self, input_folders: list, output_db_path: str):
        self.input_folders = input_folders
        self.output_db_path = output_db_path
        self.db_excel = PalkiaExcel(filepath=self.output_db_path)
        self.db_builder = DatabaseBuilder()
        self.logger = Logger()

    def _find_files(self) -> list:
        """Encontra todos os arquivos .xlsx nas pastas de entrada."""
        all_files = []
        for folder in self.input_folders:
            if not os.path.isdir(folder):
                self.logger.log(f"Pasta de entrada não encontrada: {folder}", "WARNING")
                continue
            
            files = [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith('.xlsx')]
            all_files.extend(files)
            self.logger.log(f"Encontrados {len(files)} arquivos .xlsx em '{os.path.basename(folder)}'.", "INFO")
        return all_files

    def _get_company_name_from_filename(self, filename: str) -> str:
        """Extrai um nome de empresa limpo do nome do arquivo para usar como nome da aba."""
        name = os.path.basename(filename)
        # Remove prefixos comuns e a extensão .xlsx
        name = re.sub(r'^(saida_anotacoes_|resultado_tabelas_MUST_ONS|saida_CUST-\d{4}-\d{3}-\d{2,}\s*-\s*)', '', name, flags=re.IGNORECASE)
        name = os.path.splitext(name)[0]
        # Remove sufixos comuns
        name = re.sub(r'\s*RECON.*|_minuta_recon.*', '', name, flags=re.IGNORECASE).strip()
        # Caso especial para o arquivo de tabelas unificado
        if not name or "resultado_tabelas" in os.path.basename(filename).lower():
            return "MUST_Consolidado"
        return name

    def run(self):
        """
        Executa o processo completo de construção do banco de dados.
        """
        self.logger.log("Iniciando orquestração da construção do banco de dados...", "INFO")
        files_to_process = self._find_files()

        if not files_to_process:
            self.logger.log("Nenhum arquivo .xlsx encontrado para processar. Encerrando.", "ERROR")
            return

        for filepath in files_to_process:
            filename = os.path.basename(filepath)
            self.logger.log(f"\n--- Processando arquivo: {filename} ---", "HEADER")
            
            try:
                extracted_df = pd.read_excel(filepath)
            except Exception as e:
                self.logger.log(f"ERRO ao ler o arquivo '{filename}': {e}", "ERROR")
                continue

            # A transformação só é necessária para o formato longo (anotações)
            # Para as tabelas já extraídas, podemos querer adicioná-las diretamente
            if "anotacoes" in filename.lower():
                 db_df = self.db_builder.transform_to_database_format(extracted_df)
            else:
                 # Se não for um arquivo de anotações, usamos o DataFrame como está.
                 db_df = extracted_df

            if not db_df.empty:
                company_name = self._get_company_name_from_filename(filename)
                self.db_excel.update_sheet_with_data(company_name, db_df)

        self.db_excel.save()
        self.logger.log("Processo de orquestração finalizado.", "SUCCESS")

# --- PONTO DE PARTIDA DO SCRIPT ---
if __name__ == "__main__":
    # --- 1. Configuração dos Diretórios ---
    base_folder = r"C:\Users\pedrovictor.veras\OneDrive - Operador Nacional do Sistema Eletrico\Documentos\ESTAGIO_ONS_PVRV_2025\AUTOMACÕES ONS\arquivos"
    
    # Lista de pastas onde os arquivos .xlsx de entrada estão localizados
    input_folders = [
        os.path.join(base_folder, "anotacoes_extraidas"),
        os.path.join(base_folder, "tabelas_extraidas")
    ]

    # Pasta onde o banco de dados final será salvo
    db_output_folder = os.path.join(base_folder, "database")
    os.makedirs(db_output_folder, exist_ok=True)
    
    # Caminho completo para o arquivo de banco de dados Excel
    db_filepath = os.path.join(db_output_folder, "db_consolidado.xlsx")

    # --- 2. Execução do Orquestrador ---
    orchestrator = DatabaseOrchestrator(
        input_folders=input_folders,
        output_db_path=db_filepath
    )
    orchestrator.run()

    # --- 3. Sincronização com o Banco de Dados SQLite (Opcional) ---
    # Caminho para o banco de dados de backup
    sqlite_db_path = os.path.join(db_output_folder, "database_backup.db")
    
    # Cria uma instância do builder apenas para usar os métodos de sincronização
    sync_tool = DatabaseBuilder()
    
    # Sincroniza o Excel recém-criado para o banco de dados SQLite
    sync_tool.excel_to_sql(db_filepath, sqlite_db_path)

    print("\n✅ Processo concluído com sucesso!")
