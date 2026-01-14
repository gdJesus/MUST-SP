"""
Guia Rápido de openpyxl - Equivalente a VBA
==========================================
"""
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

class ExcelOpenPyXLGuide:
    def __init__(self, filename="exemplo.xlsx"):
        self.filename = filename
        self.wb = None
        self.ws = None

    # 1. Gerenciamento de Arquivos
    def criar_arquivo(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Planilha1"
        return self.wb

    def abrir_arquivo(self):
        if os.path.exists(self.filename):
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.active
        return self.wb

    # 2. Manipulação de Planilhas
    def adicionar_planilha(self, nome):
        return self.wb.create_sheet(title=nome)

    def remover_planilha(self, nome):
        if nome in self.wb.sheetnames:
            del self.wb[nome]

    def selecionar_planilha(self, nome):
        if nome in self.wb.sheetnames:
            self.ws = self.wb[nome]
        return self.ws

    # 3. Manipulação de Células
    def escrever_celula(self, celula, valor):
        self.ws[celula] = valor

    def ler_celula(self, celula):
        return self.ws[celula].value

    def formatar_celula(self, celula, **estilo):
        cel = self.ws[celula]
        if 'fonte' in estilo:
            cel.font = Font(**estilo['fonte'])
        if 'preenchimento' in estilo:
            cel.fill = PatternFill(**estilo['preenchimento'])
        if 'alinhamento' in estilo:
            cel.alignment = Alignment(**estilo['alinhamento'])

    # 4. Gráficos
    def criar_grafico_barras(self, titulo, dados_ref, categorias_ref, posicao):
        grafico = BarChart()
        grafico.title = titulo
        grafico.add_data(dados_ref, titles_from_data=True)
        grafico.set_categories(categorias_ref)
        self.ws.add_chart(grafico, posicao)

    # 5. Operações Úteis
    def ajustar_largura_colunas(self):
        for col in self.ws.columns:
            max_length = 0
            col_letra = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ajuste = (max_length + 2) * 1.2
            self.ws.column_dimensions[col_letra].width = ajuste

    def salvar(self):
        self.wb.save(self.filename)
        return os.path.abspath(self.filename)

# Exemplo de Uso
if __name__ == "__main__":
    # 1. Criar novo arquivo
    excel = ExcelOpenPyXLGuide("exemplo.xlsx")
    excel.criar_arquivo()
    
    # 2. Adicionar dados
    dados = [
        ["Produto", "Vendas", "Meta"],
        ["Produto A", 150, 200],
        ["Produto B", 230, 200],
        ["Produto C", 180, 200],
        ["Total", "=SOMA(B2:B4)", "=SOMA(C2:C4)"]
    ]
    
    for i, linha in enumerate(dados, start=1):
        for j, valor in enumerate(linha, start=1):
            excel.ws.cell(row=i, column=j, value=valor)
    
    # 3. Formatar cabeçalho
    excel.formatar_celula("A1",
        fonte={"bold": True, "color": "FFFFFF"},
        preenchimento={"start_color": "4F81BD", "fill_type": "solid"},
        alinhamento={"horizontal": "center"}
    )
    
    # 4. Criar gráfico
    dados_ref = Reference(excel.ws, min_col=2, min_row=1, max_row=4, max_col=2)
    categorias = Reference(excel.ws, min_col=1, min_row=2, max_row=4)
    excel.criar_grafico_barras("Vendas por Produto", dados_ref, categorias, "E2")
    
    # 5. Ajustar largura das colunas
    excel.ajustar_largura_colunas()
    
    # 6. Salvar
    caminho = excel.salvar()
    print(f"Arquivo salvo em: {caminho}")
